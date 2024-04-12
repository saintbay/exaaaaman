import sqlite3
import hashlib
import requests
import datetime
from tabulate import tabulate
from openpyxl import Workbook

class User:
    def __init__(self, name):
        self.name = name
        self.registered = False
        self.purchase_history = []
        
    def is_registered(self):
        return self.registered
    
    def register(self):
        self.registered = True
    def add_to_history(self, movie_name, time, category, price):
        self.purchase_history.append({"Фильм": movie_name, "Время": time, "Категория": category, "Цена": price})
        
    def generate_receipt(self, movie_name, time, category, price):
        try:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '--- Чек ---'
            ws['A2'] = f"Имя пользователя: {self.name}"
            ws['A3'] = f"Фильм: {movie_name}"
            ws['A4'] = f"Время сеанса: {time}"
            ws['A5'] = f"Категория билета: {category}"
            ws['A6'] = f"Цена билета: {price} тг."
            ws['A7'] = "------------"
            wb.save(f"{self.name}.xlsx")
        except Exception as e:
            print(f"Ошибка при генерации чека: {e}")

class MovieManager:
    def __init__(self):
        self.connection = sqlite3.connect('database.db')
        self.cursor = self.connection.cursor()

        self.url = "https://imdb-top-100-movies.p.rapidapi.com/"
        self.headers = {
            "X-RapidAPI-Key": "731b735b0cmsh57b9761a93782bdp196090jsnd979c5b98aec",
            "X-RapidAPI-Host": "imdb-top-100-movies.p.rapidapi.com"
        }
        self.movies = []
        self.logged_in_user = None  

    def is_user_logged_in(self):
        return self.logged_in_user is not None

    def get_logged_in_user(self):
        return self.logged_in_user

    def set_logged_in_user(self, username):
        self.logged_in_user = username

    def fetch_movies(self):
        try:
            response = requests.get(self.url, headers=self.headers)
            self.movies = response.json()
        except Exception as e:
            print(f"Ошибка при загрузке фильмов: {e}")

    def get_current_showtimes(self):
        try:
            now = datetime.datetime.now()
            current_date = now.date()
            current_time = now.time()
            current_time_str = current_time.strftime("%H:%M:%S")
            self.cursor.execute("SELECT * FROM afisha WHERE date >= ? AND time > ?", (current_date, current_time_str))
            showtimes = self.cursor.fetchmany(10)  
            print("Доступные сеансы:")
            count = 0
            for showtime in showtimes:
                count += 1
                print(f"{count}. {showtime[4]} {showtime[5]}")
            return showtimes
        except Exception as e:
            print(f"Ошибка при получении текущих сеансов: {e}")
            return []

    def view_movies(self):
        try:
            print("Активные сеансы для просмотра: ")
            count = 0
            rows = []
            for movie in self.movies:
                count += 1
                if count > 12:
                    break

                name = movie.get('title', 'Unknown')
                description = movie.get('description', 'Unknown')
                rating = float(movie.get('rating', 'Unknown'))

                name = name.replace('"', "'")
                description = description.replace('"', "'")

                rows.append([count, f"{name}: {description}", rating])

            print(tabulate(rows, headers=["№", "Фильм", "Рейтинг"], tablefmt="grid"))
            verify = int(input("Введите номер фильма, который хотите посмотреть: "))
            if verify == count:
                self.cursor.execute("SELECT * FROM afisha WHERE movie_id = ?", ())
                showtimes = self.cursor.fetchall()
                showtime_rows = [[showtime[0], f"{showtime[4]} {showtime[5]}"] for showtime in showtimes]
                print(tabulate(showtime_rows, headers=["ID", "Время сеанса"], tablefmt="grid"))
        except Exception as e:
            print(f"Ошибка при просмотре фильмов: {e}")

    def choise_kinopark(self):
        try:
            self.cursor.execute("SELECT * FROM cinema")
            cinemas = self.cursor.fetchall()
            rows = []
            for cinema in cinemas:
                rows.append([cinema[0], f"{cinema[1]} ({cinema[2]})"])

            print(tabulate(rows, headers=["Номер", "Название (Адрес)"], tablefmt="grid"))

            cinema_choice = int(input("Введите номер кинопарка: "))
            return cinema_choice
        except Exception as e:
            print(f"Ошибка при выборе кинопарка: {e}")
            return None

    def buy_ticket(self, showtime_id):
        try:
            if not self.is_user_logged_in():
                print("Ошибка: пользователь не зарегистрирован. Пожалуйста, войдите в систему.")
                return

            self.cursor.execute("SELECT * FROM afisha WHERE id = ?", (showtime_id,))
            showtime = self.cursor.fetchone()
            print(f"Выбранный сеанс: {showtime[4]} {showtime[5]}")

            category = input("Выберите категорию билета (Детский, Студенческий, Взрослый): ").lower()
            if category not in ["детский", "студенческий", "взрослый"]:
                print("Ошибка: неверная категория билета.")
                return

            price = showtime[3]
            if category == "детский":
                price = int(price * 0.3)
            elif category == "студенческий":
                price = int(price * 0.5)

            print(f"Цена билета: {price} тг.")
            confirm = input("Желаете купить билет? (да/нет): ").lower()
            if confirm == "да":
                self.cursor.execute("INSERT INTO ticket(name, phone, place_id) VALUES(?, ?, ?)", ("", "", showtime_id))
                self.connection.commit()
                print("Билет куплен успешно!")

                user = User(self.logged_in_user)
                user.generate_receipt(showtime[4], showtime[5], category, price)
            else:
                print("Покупка отменена.")
        except Exception as e:
            print(f"Ошибка при покупке билета: {e}")

    def register_user(self):
        try:
            print("Регистрация нового пользователя:")
            name = input("Введите ваше имя: ")
            card_number = input("Введите номер банковской карты: ")
            password = input("Введите пароль: ")

            password_hash = hashlib.md5(password.encode()).hexdigest()

            self.cursor.execute("INSERT INTO users (name, card_number, password_hash) VALUES (?, ?, ?)",
                                (name, card_number, password_hash))
            self.connection.commit()

            print("Регистрация завершена успешно.")
        except Exception as e:
            print(f"Ошибка при регистрации пользователя: {e}")

    def login_user(self):
        try:
            print("Вход в систему:")
            card_number = input("Введите номер банковской карты: ")
            password = input("Введите пароль: ")

            password_hash = hashlib.md5(password.encode()).hexdigest()

            self.cursor.execute("SELECT * FROM users WHERE card_number = ? AND password_hash = ?",
                                (card_number, password_hash))
            user_data = self.cursor.fetchone()
            if user_data:
                print(f"Добро пожаловать {user_data[1]}!")
                self.set_logged_in_user(user_data[1])
            else:
                print("Ошибка входа проверьте номер карты и пароль.")
        except Exception as e:
            print(f"Ошибка при входе в систему: {e}")

    def run(self):
        print("Здравствуйте! Выберите что хотите сделать:")
        while True:
            try:
                choice = int(input(" 1) Посмотреть фильм \n 2) Регистрация \n 3) Вход в учетную запись \n 4) Выход \n"))
                if choice == 1:
                    self.fetch_movies()
                    self.view_movies()
                    cinema_choice = self.choise_kinopark()
                    if cinema_choice is not None:
                        showtimes = self.get_current_showtimes()
                        if showtimes:
                            showtime_choice = int(input("Введите номер удобного сеанса: "))
                            print("Время сеанса:", showtime_choice)
                            showtime_id = showtimes[showtime_choice - 1][0]
                            self.buy_ticket(showtime_id)

                elif choice == 2:
                    self.register_user()
                elif choice == 3:
                    self.login_user()
                elif choice == 4:
                    break
                else:
                    print("Неверный выбор попробуйте еще раз")
            except ValueError:
                print("Неверный ввод пожалуйста введите число")

movie_manager = MovieManager()
movie_manager.run()


# import requests
# import json
# import sqlite3
# import random
# import datetime

# connection = sqlite3.connect('database.db')
# cursor = connection.cursor()

# url = "https://imdb-top-100-movies.p.rapidapi.com/"

# headers = {
# 	"X-RapidAPI-Key": "731b735b0cmsh57b9761a93782bdp196090jsnd979c5b98aec",
# 	"X-RapidAPI-Host": "imdb-top-100-movies.p.rapidapi.com"
# }

# response = requests.get(url, headers=headers)

# movies = response.json()

# for movie in movies:
#     print(movie.get('title', 'Unknown'))
    
#     name = movie.get('title', 'Unknown')
#     genre = movie.get('genre', 'Unknown')[0]
#     year = movie.get('year', 'Unknown')
#     description = movie.get('description', 'Unknown')
#     rating = float(movie.get('rating', 'Unknown'))
    
#     name = name.replace('"', "'")
#     genre = genre.replace('"', "'")
#     description = description.replace('"', "'")
    
#     print(name, genre, year, description, rating)
    
#     cursor.execute(f'INSERT INTO movie (name, genre, year, description, rating) VALUES ("{name}", "{genre}", {year}, "{description}", {rating})')
#     connection.commit()


# cinemas = [
#     {
#         'name' : 'Арман',
#         'address' : '​Проспект Кабанбай батыр, 21'  
#     },
#     {
#         'name' : 'Chaplin cinemas',
#         'address' : '​Проспект Туран, 37'  
#     },
#     {
#         'name' : 'Kinopark ',
#         'address' : '​Проспект Туран, 24'  
#     },
#     {
#         'name' : 'Евразия Cinema 7',
#         'address' : '​Улица Алексея Петрова, 24а/1'  
#     },
#     {
#         'name' : 'Арсенал ',
#         'address' : '​Улица Ыбырай Алтынсарин, 4'  
#     },
# ]

# for cinema in cinemas:
#     cursor.execute(f'INSERT INTO cinema (name, address) VALUES ("{cinema.get("name")}", "{cinema.get("address")}")')
#     connection.commit()


# start_date = datetime.date.today()
# end_date = datetime.date(2024, 5, 30)
# start_time = datetime.time(16, 0, 0)
# end_time = datetime.time(0, 0, 0)
# movies_id = [random.randint(1, 100) for i in range(1, 51)]
# cinemas_id = [random.randint(1, 5) for i in range(1, 51)]
# prices = [random.randint(1000, 5000) for i in range(1, 51)]
# gap = int((end_date - start_date).total_seconds())
# dates = [start_date + datetime.timedelta(seconds=random.randint(0, gap)) for i in range(1, 51)]
# times = [f"{random.randint(16, 23)}:{random.randint(0, 5)}0:00" for i in range(1, 51)]
# capacities = [random.randint(50, 100) for i in range(1, 51)]
# for i in range(0, 50):
#     cursor.execute(
#         f'INSERT INTO afisha (movie_id, cinema_id, price, date, time, capacity) VALUES ({movies_id[i]}, {cinemas_id[i]}, {prices[i]}, "{dates[i]}", "{times[i]}", {capacities[i]})'
#     )
#     connection.commit()

# afishas = [i for i in range(1, 51)]
# rooms = [random.randint(1, 6) for i in range(1, 51)] 
# rows = [random.randint(1, 16) for i in range(1, 51)]
# columns = [random.randint(1, 16) for i in range(1, 51)]

# for i in range(0, 50):
#     cursor.execute(
#         f"INSERT INTO place (afisha_id, room, row, seat) VALUES ({afishas[i]}, {rooms[i]}, {rows[i]}, {columns[i]})"
#     )
#     connection.commit()